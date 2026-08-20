import ast
import time
import pandas as pd
from datetime import datetime
from re import findall

from PyQt5.QtCore import Qt, QDir, QTimer, QEvent
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import (QApplication, QMainWindow, QGraphicsView,
                             QFileDialog, QTableWidgetItem, QComboBox, QLabel)
from main_window.ui2 import UI
from chip.DMF_Chip import DMF_Chip
from serial_driver.Serialdriver import SerialDriver
from cam.cam_module import CameraLogic
from logger.log import BrowseLog
from utils.path_helper import get_resource_path
import pyqtgraph as pg

CAM_ENABLE = 0


class PlotZoomWindow(QMainWindow):
    """独立放大窗口：复制曲线数据，不移动原 Tab 中的 PlotWidget。"""
    def __init__(self, plot_key, main_window):
        super().__init__(None)
        self._plot_key = plot_key
        self._main_window = main_window
        if plot_key == "fluor":
            self.setWindowTitle("荧光曲线（双击关闭）")
        else:
            self.setWindowTitle(f"通道{plot_key} 温度曲线（双击关闭）")
        self.setWindowFlags(Qt.Window)
        self.resize(1200, 800)

        self._plot = pg.PlotWidget()
        self._plot.setBackground("#f7fafc")
        self._plot.setLabel("bottom", "时间 (s)")
        self._plot.showGrid(x=True, y=True, alpha=0.15)
        plot_item = self._plot.getPlotItem()
        plot_item.getViewBox().setLimits(xMin=0)
        plot_item.enableAutoRange(axis="x", enable=True)
        plot_item.enableAutoRange(axis="y", enable=True)

        self._temp_curve = None
        self._power_curve = None
        self._power_vb = None
        if plot_key == "fluor":
            left = plot_item.getAxis("left")
            left.setPen(pg.mkPen(color=(128, 0, 128), width=1))
            left.setTextPen(pg.mkPen(color=(128, 0, 128)))
            left.setLabel("荧光强度", color=(128, 0, 128))
            self._temp_curve = self._plot.plot(
                [], [], pen=pg.mkPen(color=(128, 0, 128), width=2)
            )
        else:
            left = plot_item.getAxis("left")
            left.setPen(pg.mkPen(color="red", width=1))
            left.setTextPen(pg.mkPen(color="red"))
            left.setLabel("温度(°C)", color="red")
            right = plot_item.getAxis("right")
            right.setPen(pg.mkPen(color="green", width=1))
            right.setTextPen(pg.mkPen(color="green"))
            right.setLabel("功率(W)", color="green")
            plot_item.showAxis("right", True)
            self._temp_curve = self._plot.plot([], [], pen=pg.mkPen(color="red", width=2))
            self._power_vb = pg.ViewBox()
            plot_item.scene().addItem(self._power_vb)
            right.linkToView(self._power_vb)
            plot_item.getViewBox().sigXRangeChanged.connect(
                lambda: self._power_vb.setXRange(*plot_item.getViewBox().viewRange()[0])
            )
            self._power_vb.setYRange(0, 5, padding=0)
            self._power_curve = pg.PlotDataItem([], [], pen=pg.mkPen(color="green", width=2))
            self._power_vb.addItem(self._power_curve)
            plot_item.layout.addItem(self._power_vb, 2, 1)

        self.setCentralWidget(self._plot)
        self._sync_timer = QTimer(self)
        self._sync_timer.setInterval(200)
        self._sync_timer.timeout.connect(self._sync_from_main)
        self._hook_plot_double_click(self._plot, self.close)

    @staticmethod
    def _hook_plot_double_click(plot_widget, on_double_click):
        vb = plot_widget.getPlotItem().getViewBox()
        orig_click = vb.mouseClickEvent

        def mouseClickEvent(ev):
            if ev.double():
                on_double_click()
            if orig_click is not None:
                orig_click(ev)

        vb.mouseClickEvent = mouseClickEvent

    def _sync_from_main(self):
        mw = self._main_window
        key = self._plot_key
        if key == "fluor":
            self._temp_curve.setData(mw.fluor_time, mw.fluor_data)
        elif key == 1:
            self._temp_curve.setData(mw.temp1_time, mw.temp1_data)
            if self._power_curve and mw.power1_data:
                self._power_curve.setData(mw.temp1_time, mw.power1_data)
        elif key == 2:
            self._temp_curve.setData(mw.temp2_time, mw.temp2_data)
            if self._power_curve and mw.power2_data:
                self._power_curve.setData(mw.temp2_time, mw.power2_data)
        elif key == 3:
            self._temp_curve.setData(mw.temp3_time, mw.temp3_data)
            if self._power_curve and mw.power3_data:
                self._power_curve.setData(mw.temp3_time, mw.power3_data)
        if self._temp_curve.xData is not None and len(self._temp_curve.xData) > 0:
            self._plot.getPlotItem().getViewBox().autoRange(padding=0.02)

    def showEvent(self, event):
        super().showEvent(event)
        self._sync_from_main()
        self._sync_timer.start()

    def closeEvent(self, event):
        self._sync_timer.stop()
        mw = self._main_window
        if mw is not None and mw._plot_zoom_windows.get(self._plot_key) is self:
            mw._plot_zoom_windows.pop(self._plot_key, None)
        event.accept()
class CamZoomWindow(QMainWindow):
    """独立摄像头放大窗口：同步主窗口当前帧，不移动原 QLabel。"""
    def __init__(self, main_window):
        super().__init__(None)
        self._main_window = main_window
        self.setWindowTitle("摄像头（双击关闭）")
        self.setWindowFlags(Qt.Window)
        self.resize(1200, 800)
        self._label = QLabel()
        self._label.setAlignment(Qt.AlignCenter)
        self._label.setStyleSheet("background: #0a1e30; color: #b9d1e4; font-size: 48px; font-weight: bold;")
        self._label.setText("CAM")
        self.setCentralWidget(self._label)
        self._sync_timer = QTimer(self)
        self._sync_timer.setInterval(40)
        self._sync_timer.timeout.connect(self._sync_from_main)
        def mouseDoubleClickEvent(event):
            self.close()
            event.accept()
        self._label.mouseDoubleClickEvent = mouseDoubleClickEvent
    def _sync_from_main(self):
        mw = self._main_window
        camera_logic = getattr(mw, "camera_logic", None)
        if camera_logic is None:
            return
        qimage = camera_logic.current_qimage
        if qimage is None or qimage.isNull():
            if not camera_logic.timer_camera.isActive():
                self._label.clear()
                self._label.setText("CAM")
            return
        label_size = self._label.size()
        if label_size.width() < 10 or label_size.height() < 10:
            return
        scaled = QPixmap.fromImage(qimage).scaled(
            label_size.width(),
            label_size.height(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        self._label.setPixmap(scaled)
    def showEvent(self, event):
        super().showEvent(event)
        self._sync_from_main()
        self._sync_timer.start()
    def closeEvent(self, event):
        self._sync_timer.stop()
        mw = self._main_window
        if mw is not None and mw._cam_zoom_window is self:
            mw._cam_zoom_window = None
        event.accept()

class MainWindow(QMainWindow, UI):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.SetUpUi(self)
        chip_csv_path = get_resource_path('chip/chip_point.csv')
        self.dmf_chip_scene = DMF_Chip(chip_csv_path)
        self.log=BrowseLog(text_browser1=self.log_information,text_browser2=self.serial_information)
        self.driver = SerialDriver(log=self.log)
        self.graphicsView1.setScene(self.dmf_chip_scene)
        self.graphicsView1.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.graphicsView1.setDragMode(QGraphicsView.RubberBandDrag)
        self.graphicsView1.setCacheMode(QGraphicsView.CacheBackground)

        """串口设置"""
        self.button_flush_serial_port.clicked.connect(self.flush_port)
        self.button_open_serial_port.clicked.connect(self.open_port)
        self.button_close_serial_port.clicked.connect(self.close_port)
        self.button1.clicked.connect(self.phy_link)
        self.button2.clicked.connect(self.left_link)
        self.button3.clicked.connect(self.right_link)
        """参数设置"""
        self.lineEdit1.returnPressed.connect(self.set_voltage)
        self.lineEdit2.returnPressed.connect(self.set_acdc)
        self.lineEdit3_1.returnPressed.connect(self.set_times)
        self.lineEdit3_2.returnPressed.connect(self.set_times1)
        """定时器用于按时发送驱动数据"""
        self.timer = QTimer(self)
        self.timer.setInterval(1000)
        self.timer.timeout.connect(self.timer_send_data)
        """手动操作界面函数绑定"""
        self.button_up.clicked.connect(self.move_up)
        self.button_down.clicked.connect(self.move_down)
        self.button_left.clicked.connect(self.move_left)
        self.button_right.clicked.connect(self.move_right)
        self.button_zoomout.clicked.connect(self.zoomout)
        self.button_zoomin.clicked.connect(self.zoomin)
        self.dmf_chip_scene.serialDataSignal.connect(self._send_drive_bitmap)
        """自动操作界面槽函数绑定 """
        self.button_add.clicked.connect(self.add_data)
        self.button_del.clicked.connect(self.del_data)
        self.button_sta.clicked.connect(self.drive_start)
        self.button_end.clicked.connect(self.drive_end)
        self.button_save_data.clicked.connect(self.save_data)
        self.button_load_data.clicked.connect(self.load_data)
        """液滴位置检测"""
        self.button_detect_droplet.clicked.connect(self.detect_droplet)
        """摄像头模块初始化"""
        if CAM_ENABLE:
            self.camera_logic = CameraLogic(
                label_view=self.label_cam_view,
                button_open=self.button_open_camera,
                button_close=self.button_close_camera,
                button_take_picture=self.button_take_picture,
                button_record_picture=self.button_record_picture,
                logger=self.log,
                cam_number=1,
            )
        """温度曲线绘制初始化"""
        # 初始化温度数据缓冲区，存储时间和数值
        # 通道1（左侧）: 温度和功率
        self.temp1_time = []  # 时间序列
        self.temp1_data = []  # 温度数据
        self.power1_data = []  # 功率数据
        # 通道2（右侧）: 温度和功率
        self.temp2_time = []
        self.temp2_data = []
        self.power2_data = []
        # 通道3（如果有）: 温度和功率
        self.temp3_time = []
        self.temp3_data = []
        self.power3_data = []

        self.max_data_points = 12000  # 最大数据点数，防止内存溢出
        self.temp_start_time = {}  # 记录每个通道的开始时间 {channel: start_time}

        # 为每个通道创建独立的温度读取定时器
        self.temp_timers = {}
        self.temp_timers[1] = QTimer(self)
        self.temp_timers[1].setInterval(100)  # 100ms间隔，1秒读取10次
        self.temp_timers[1].timeout.connect(lambda: self.read_temperature_and_plot(1))
        self.temp_timers[2] = QTimer(self)
        self.temp_timers[2].setInterval(100)  # 100ms间隔，1秒读取10次
        self.temp_timers[2].timeout.connect(lambda: self.read_temperature_and_plot(2))
        self.temp_timers[3] = QTimer(self)
        self.temp_timers[3].setInterval(100)  # 100ms间隔，1秒读取10次
        self.temp_timers[3].timeout.connect(lambda: self.read_temperature_and_plot(3))

        # 绑定温度控制按钮
        self.button_start1.clicked.connect(lambda: self.start_temp_monitoring(1))
        self.button_pause1.clicked.connect(lambda: self.pause_temp_monitoring(1))
        self.button_start2.clicked.connect(lambda: self.start_temp_monitoring(2))
        self.button_pause2.clicked.connect(lambda: self.pause_temp_monitoring(2))
        self.button_start3.clicked.connect(lambda: self.start_temp_monitoring(3))
        self.button_pause3.clicked.connect(lambda: self.pause_temp_monitoring(3))

        # 绑定PID参数输入框的Enter键事件
        self.lineEdit_P1.returnPressed.connect(lambda: self.send_pid_params(1))
        self.lineEdit_I1.returnPressed.connect(lambda: self.send_pid_params(1))
        self.lineEdit_D1.returnPressed.connect(lambda: self.send_pid_params(1))
        
        self.lineEdit_P2.returnPressed.connect(lambda: self.send_pid_params(2))
        self.lineEdit_I2.returnPressed.connect(lambda: self.send_pid_params(2))
        self.lineEdit_D2.returnPressed.connect(lambda: self.send_pid_params(2))
        
        self.lineEdit_P3.returnPressed.connect(lambda: self.send_pid_params(3))
        self.lineEdit_I3.returnPressed.connect(lambda: self.send_pid_params(3))
        self.lineEdit_D3.returnPressed.connect(lambda: self.send_pid_params(3))

        # 绑定温度数据保存和加载按钮
        self.button_save1.clicked.connect(lambda: self.save_temp_data(1))
        self.button_load1.clicked.connect(lambda: self.load_temp_data(1))
        self.button_clear1.clicked.connect(lambda: self.clear_temp_data(1))
        self.button_save2.clicked.connect(lambda: self.save_temp_data(2))
        self.button_load2.clicked.connect(lambda: self.load_temp_data(2))
        self.button_clear2.clicked.connect(lambda: self.clear_temp_data(2))
        self.button_save3.clicked.connect(lambda: self.save_temp_data(3))
        self.button_load3.clicked.connect(lambda: self.load_temp_data(3))
        self.button_clear3.clicked.connect(lambda: self.clear_temp_data(3))

        #绑定荧光读取相关
        self.fluor_time = []
        self.fluor_data = []
        self.fluor_start_time = None
        self.fluor_excitation_time = 5
        self.fluor_interval_ms = 1000
        self._fluor_led_on = False
        self._fluor_method1_on = False
        self._fluor_method2_on = False
        self.fluor_timer = QTimer(self)
        self.fluor_timer.timeout.connect(self._fluor_timer_tick)
        self.button_fluor_led.clicked.connect(self.toggle_fluor_led)
        self.button_fluor_method1.clicked.connect(self.toggle_fluor_method1)
        self.button_fluor_method2.clicked.connect(self.toggle_fluor_method2)
        self.button_fluor_save.clicked.connect(self.save_fluor_data)
        self.button_fluor_load.clicked.connect(self.load_fluor_data)
        self.button_fluor_clear.clicked.connect(self.clear_fluor_data)
        self._plot_zoom_windows = {}
        self._cam_zoom_window = None
        self._init_plot_double_click()
        self._init_cam_double_click()
        QApplication.instance().installEventFilter(self)
    #摄像头框双击放大相关
    def _init_cam_double_click(self):
        if not CAM_ENABLE:
            return
        self.label_cam_view.mouseDoubleClickEvent = self._on_cam_view_double_click
    def _on_cam_view_double_click(self, event):
        self._toggle_cam_fullscreen()
        event.accept()
    def _toggle_cam_fullscreen(self):
        if self._cam_zoom_window is not None:
            self._cam_zoom_window.close()
            return
        if not CAM_ENABLE:
            return
        try:
            window = CamZoomWindow(self)
            self._cam_zoom_window = window
            window.show()
            window.showMaximized()
            window.raise_()
            window.activateWindow()
        except Exception as e:
            self._cam_zoom_window = None
            self.log.debug(False, f"摄像头放大失败: {e}")
    #温度窗口和荧光窗口放大相关
    def _init_plot_double_click(self):
        for plot_key, plot_widget in (
            (1, self.temperature_widget1),
            (2, self.temperature_widget2),
            (3, self.temperature_widget3),
            ("fluor", self.fluorescence_widget),
        ):
            PlotZoomWindow._hook_plot_double_click(
                plot_widget,
                lambda key=plot_key: self._toggle_plot_fullscreen(key),
            )
    def _toggle_plot_fullscreen(self, plot_key):
        if plot_key in self._plot_zoom_windows:
            self._plot_zoom_windows[plot_key].close()
            return
        try:
            window = PlotZoomWindow(plot_key, self)
            self._plot_zoom_windows[plot_key] = window
            window.show()
            window.showMaximized()
            window.raise_()
            window.activateWindow()
        except Exception as e:
            self._plot_zoom_windows.pop(plot_key, None)
            self.log.debug(False, f"放大失败: {e}")
    def _active_fluor_driver(self):
        if self.driver.ser is not None and self.driver.ser.isOpen():
            return self.driver
        return None
    def _update_fluor_led_button(self):
        self.button_fluor_led.setText("关闭LED" if self._fluor_led_on else "打开LED")
    def toggle_fluor_led(self):
        driver = self._active_fluor_driver()
        if driver is None:
            self.log.debug(False, "请先在左侧「系统初始化」中打开串口")
            return
        try:
            if self._fluor_led_on:
                driver.detector_cmd2()
                self._fluor_led_on = False
                self.log.debug(True, "激发 LED 已关闭")
            else:
                driver.detector_cmd3()
                self._fluor_led_on = True
                self.log.debug(True, "激发 LED 已打开")
            self._update_fluor_led_button()
        except Exception as e:
            action = "关闭" if self._fluor_led_on else "打开"
            self.log.debug(False, f"{action}激发 LED 失败: {e}")
    def _refresh_fluor_input_vars(self):
        exc_text = self.lineEdit_fluor_excitation.text().strip()
        interval_text = self.lineEdit_fluor_interval.text().strip()
        try:
            exc_val = int(exc_text) if exc_text else 20
        except ValueError:
            exc_val = 20
        self.fluor_excitation_time = max(1, min(200, exc_val))

        try:
            interval_val = int(interval_text) if interval_text else 500
        except ValueError:
            interval_val = 500
        self.fluor_interval_ms = max(1, min(1000, interval_val))
    def _update_fluor_method1_button(self):
        self.button_fluor_method1.setText("暂停①" if self._fluor_method1_on else "开始①")
    def _update_fluor_method2_button(self):
        self.button_fluor_method2.setText("暂停②" if self._fluor_method2_on else "开始②")
    def _fluor_timer_tick(self):
        if self._fluor_method1_on:
            self._fluor_method1_tick()
        elif self._fluor_method2_on:
            self._fluor_method2_tick()
    def _fluor_method1_tick(self):
        self._refresh_fluor_input_vars()
        driver = self._active_fluor_driver()
        if driver is None:
            self.log.debug(False, "请先在左侧「系统初始化」中打开串口")
            self.fluor_timer.stop()
            self._fluor_method1_on = False
            self._update_fluor_method1_button()
            return
        try:
            value = driver.detector_cmd1(self.fluor_excitation_time)
            if value is None:
                self.log.debug(False, "方法一荧光读取无响应")
                return
            # if value >= 16777215:
            #     self.log.debug(False, "方法一荧光读取数值无效")
            #     return
            self._append_fluor_point(value)
            self.log.debug(True, f"方法一荧光采样: {value}")
        except Exception as e:
            self.log.debug(False, f"方法一荧光采样失败: {e}")
            self.fluor_timer.stop()
            self._fluor_method1_on = False
            self._update_fluor_method1_button()
    def toggle_fluor_method1(self):
        if self._fluor_method1_on:
            self.fluor_timer.stop()
            self._fluor_method1_on = False
            self._update_fluor_method1_button()
            self.log.debug(True, "方法一荧光检测已暂停")
            return
        if self._active_fluor_driver() is None:
            self.log.debug(False, "请先在左侧「系统初始化」中打开串口")
            return
        if self._fluor_method2_on:
            self.log.debug(False, "方法二正在运行，请先暂停方法二")
            return
        self._refresh_fluor_input_vars()
        if self.fluor_start_time is None:
            self.fluor_start_time = time.time()
        self._fluor_method1_on = True
        self._update_fluor_method1_button()
        self.fluor_timer.setInterval(self.fluor_interval_ms)
        self._fluor_method1_tick()
        self.fluor_timer.start()
        self.log.debug(True, f"方法一荧光检测已开始 ({self.fluor_interval_ms}ms)")
    def _read_fluor_value_once(self):
        driver = self._active_fluor_driver()
        if driver is None:
            self.log.debug(False, "请先在左侧「系统初始化」中打开串口")
            return False
        value = driver.detector_cmd4()
        if value is None:
            self.log.debug(False, "荧光读取无响应")
            return False
        if value >= 16777215:
            self.log.debug(False, "荧光读取数值无效")
            return False
        self._append_fluor_point(value)
        return True
    def toggle_fluor_method2(self):
        if self._fluor_method2_on:
            self.fluor_timer.stop()
            self._fluor_method2_on = False
            self._update_fluor_method2_button()
            self.log.debug(True, "方法二荧光曲线采样已暂停")
            return
        if self._active_fluor_driver() is None:
            self.log.debug(False, "请先在左侧「系统初始化」中打开串口")
            return
        if self._fluor_method1_on:
            self.log.debug(False, "方法一正在运行，请先暂停方法一")
            return
        self._refresh_fluor_input_vars()
        if self.fluor_start_time is None:
            self.fluor_start_time = time.time()
        self._fluor_method2_on = True
        self._update_fluor_method2_button()
        self.fluor_timer.setInterval(self.fluor_interval_ms)
        self.fluor_timer.start()
        self.log.debug(True, f"方法二荧光曲线采样已开始 ({self.fluor_interval_ms}ms)")
    def _fluor_method2_tick(self):
        try:
            self._read_fluor_value_once()
        except Exception as e:
            self.log.debug(False, f"方法二荧光采样失败: {e}")
            self.fluor_timer.stop()
            self._fluor_method2_on = False
            self._update_fluor_method2_button()
    def _append_fluor_point(self, value):
        if self.fluor_start_time is None:
            self.fluor_start_time = time.time()
        t = time.time() - self.fluor_start_time
        self.fluor_time.append(t)
        self.fluor_data.append(float(value))
        if len(self.fluor_time) > self.max_data_points:
            self.fluor_time = self.fluor_time[-self.max_data_points:]
            self.fluor_data = self.fluor_data[-self.max_data_points:]
        self.fluorescence_curve.setData(self.fluor_time, self.fluor_data)
        self.fluorescence_widget.getPlotItem().getViewBox().autoRange(padding=0.02)
    def clear_fluor_data(self):
        self.fluor_timer.stop()
        self._fluor_method1_on = False
        self._fluor_method2_on = False
        self._update_fluor_method1_button()
        self._update_fluor_method2_button()
        self.fluor_time = []
        self.fluor_data = []
        self.fluor_start_time = None
        self.fluorescence_curve.setData([], [])
        self.log.debug(True, "已清除荧光曲线数据")
    def save_fluor_data(self):
        if not self.fluor_time:
            self.log.debug(False, "没有荧光数据可保存")
            return
        try:
            dlg = QFileDialog()
            dlg.setFileMode(QFileDialog.AnyFile)
            dlg.setNameFilter("Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;所有文件 (*)")
            dlg.setFilter(QDir.Files | QDir.NoDotAndDotDot)
            timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
            dlg.setWindowTitle("保存荧光数据")
            dlg.setAcceptMode(QFileDialog.AcceptSave)
            dlg.setDefaultSuffix("xlsx")
            dlg.selectFile(f"荧光数据_{timestamp}.xlsx")
            if not dlg.exec_():
                return
            file_path = dlg.selectedFiles()[0]
            if file_path.lower().endswith(".csv"):
                df = pd.DataFrame({"时间(s)": self.fluor_time, "荧光强度": self.fluor_data})
                df.to_csv(file_path, index=False, encoding="utf-8-sig")
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "荧光数据"
                ws.append(["时间(s)", "相对荧光强度"])
                for t, v in zip(self.fluor_time, self.fluor_data):
                    ws.append([t, v])
                wb.save(file_path)
            self.log.debug(True, f"荧光数据已保存到 {file_path}")
        except Exception as e:
            self.log.debug(False, f"保存荧光数据失败: {e}")
    def load_fluor_data(self):
        try:
            dlg = QFileDialog()
            dlg.setFileMode(QFileDialog.ExistingFile)
            dlg.setNameFilter("Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;所有文件 (*)")
            dlg.setWindowTitle("加载荧光数据")
            if not dlg.exec_():
                return
            file_path = dlg.selectedFiles()[0]
            if file_path.lower().endswith(".csv"):
                df = pd.read_csv(file_path)
                time_col = df.columns[0]
                value_col = df.columns[1]
                times = df[time_col].tolist()
                values = df[value_col].tolist()
            else:
                df = pd.read_excel(file_path)
                times = df.iloc[:, 0].tolist()
                values = df.iloc[:, 1].tolist()
            self.fluor_time = [float(x) for x in times]
            self.fluor_data = [float(x) for x in values]
            self.fluor_start_time = time.time() - (self.fluor_time[-1] if self.fluor_time else 0)
            self.fluorescence_curve.setData(self.fluor_time, self.fluor_data)
            self.fluorescence_widget.getPlotItem().getViewBox().autoRange(padding=0.02)
            self.log.debug(True, f"荧光数据已加载: {file_path}")
        except Exception as e:
            self.log.debug(False, f"加载荧光数据失败: {e}")
    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyRelease and event.key() == Qt.Key_Control:
            self.dmf_chip_scene.flush_ctrl_right_pending()

        if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Escape:
            if self._plot_zoom_windows or self._cam_zoom_window is not None:
                for win in list(self._plot_zoom_windows.values()):
                    win.close()
                if self._cam_zoom_window is not None:
                    self._cam_zoom_window.close()
                return True

        return super().eventFilter(obj, event)
    def closeEvent(self, event):
        for win in list(self._plot_zoom_windows.values()):
            win.close()
        self._plot_zoom_windows.clear()
        if self._cam_zoom_window is not None:
            self._cam_zoom_window.close()
            self._cam_zoom_window = None
        self.fluor_timer.stop()
        super().closeEvent(event)

    """串口初始化槽函数"""
    def open_port(self):
        if self.driver.open(findall(r'COM\d+',self.comboBox.currentText())[0]):
            self.button_open_serial_port.setEnabled(False)
            self._set_header_status("串口已连接", "connected")
    def flush_port(self):
        self.button_open_serial_port.setEnabled(True)
        self.button_close_serial_port.setEnabled(True)
        self.comboBox.clear()
        self.comboBox.addItems(self.driver.find_available_port())
        self._set_header_status("端口列表已刷新", "notice")
    def close_port(self):
        if self.driver.close():
            self.button_close_serial_port.setEnabled(False)
            self._set_header_status("设备未连接", "notice")

    def _set_header_status(self, text, state):
        """Update the compact connection indicator in the header."""
        self.header_status.setText(text)
        self.header_status.setProperty("state", state)
        self.header_status.style().unpolish(self.header_status)
        self.header_status.style().polish(self.header_status)

    @staticmethod
    def _set_link_status(label, healthy):
        label.setProperty("state", "ok" if healthy else "error")
        label.style().unpolish(label)
        label.style().polish(label)
    def phy_link(self):
        self._set_link_status(self.label1_button, self.driver.bottom_cmd1() == 0)
    def left_link(self):
        self._set_link_status(self.label2_button, self.driver.temp_cmd1(0x00) == 0)
    def right_link(self):
        self._set_link_status(self.label3_button, self.driver.temp_cmd2(0x00) == 0)
    """手动操控界面槽函数"""
    def _driving_electrodes(self):
        return [e for e in self.dmf_chip_scene._iter_electrodes() if e.isDrive and not e.isPinned]
    def _send_drive_bitmap(self):
        data = self.dmf_chip_scene.data_transfer(self.dmf_chip_scene.hardware_drive_ids())
        if sum(data) != 0:
            self.driver.drive_cmd6(data)
    def _move_droplets(self, dr, dc):
        """界面上下左右：移动已驱动液滴到邻接电极，再下发完整驱动位图（固定电极不移动）。"""
        sources = self._driving_electrodes()
        if not sources:
            if self.dmf_chip_scene.pinned_ids():
                self.log.debug(True, "固定电极不参与移动；请先右键驱动其它液滴")
            else:
                self.log.debug(True, "请先右键驱动液滴后再移动")
            return

        arr = self.dmf_chip_scene.id_array
        pos_map = self.dmf_chip_scene.id_position
        source_ids = {src.ID for src in sources}
        candidates = []
        blocked = []

        for src in sources:
            position = pos_map.get(src.ID)
            if position is None:
                blocked.append(src)
                continue
            r, c = position[0] + dr, position[1] + dc
            if not (0 <= r < len(arr) and 0 <= c < len(arr[0])):
                blocked.append(src)
                continue
            nid = arr[r][c]
            if nid == 0:
                blocked.append(src)
                continue
            target = self.dmf_chip_scene._electrode_by_id(nid)
            if target is not None and target.isPinned:
                blocked.append(src)
                continue
            candidates.append((nid, src))

        by_target = {}
        for nid, src in candidates:
            if nid in by_target:
                blocked.append(src)
                blocked.append(by_target[nid])
                del by_target[nid]
            else:
                by_target[nid] = src

        for nid, src in list(by_target.items()):
            occupant = self.dmf_chip_scene._electrode_by_id(nid)
            if (
                occupant is not None
                and occupant.isDrive
                and occupant.ID not in source_ids
                and not occupant.isPinned
            ):
                blocked.append(src)
                del by_target[nid]

        blocked_ids = {src.ID for src in blocked}
        for src in blocked:
            self.dmf_chip_scene.set_electrode_state(src.ID, drive=False, selected=False)

        if blocked and not by_target:
            self.log.debug(True, "已到达边界：液滴消失（未选中、未驱动）")

        if not by_target:
            if blocked_ids or blocked:
                self._send_drive_bitmap()
            return

        for src in by_target.values():
            self.dmf_chip_scene.set_electrode_state(src.ID, drive=False, selected=False)
        for nid in by_target:
            self.dmf_chip_scene.set_electrode_state(nid, drive=True, selected=True)

        self._send_drive_bitmap()
    def move_up(self):
        self._move_droplets(-1, 0)
    def move_down(self):
        self._move_droplets(1, 0)
    def move_left(self):
        self._move_droplets(0, -1)
    def move_right(self):
        self._move_droplets(0, 1)
    def zoomout(self):
        if self.dmf_chip_scene.scale_now<=1.4:
            self.graphicsView1.scale(1.1, 1.1)
            self.dmf_chip_scene.scale_now*=1.1
    def zoomin(self):
        if self.dmf_chip_scene.scale_now >= 0.7:
            self.graphicsView1.scale(0.9, 0.9)
            self.dmf_chip_scene.scale_now *= 0.9
    """自动操作界面函数"""
    def add_data(self):
        current_row = self.tab2_tab.currentRow()
        if current_row == -1:
            insert_row = self.tab2_tab.rowCount()
        else:
            insert_row = current_row + 1
        self.add_one_row(frame_list=[], drive_time=1, camera_enable="关", insert_row=insert_row)
    def del_data(self):
        selected_row = self.tab2_tab.selectedItems()
        if selected_row:
            row_to_delete = selected_row[0].row()
        else:
            row_count = self.tab2_tab.rowCount()
            if row_count > 0:
                row_to_delete = row_count - 1
            else:
                return
        
        self.tab2_tab.removeRow(row_to_delete)
        # 更新行号
        for i in range(self.tab2_tab.rowCount()):
            self.tab2_tab.verticalHeaderItem(i).setText(str(i + 1))
    def drive_start(self):
        if self.tab2_tab.rowCount() >= 1:
            if self.tab2_tab.currentRow() == -1:
                self.tab2_tab.setCurrentCell(0, 0)
                self.tab2_tab.scrollToItem(self.tab2_tab.item(0, 0))

            drive_time_item = self.tab2_tab.item(0, 1)
            drive_time = 1
            if drive_time_item and drive_time_item.text().strip():
                try:
                    drive_time = int(drive_time_item.text().strip())
                    if drive_time < 1:
                        drive_time = 1
                except:
                    drive_time = 1
            self.timer.setInterval(drive_time * 1000)
            self.timer_send_data()
        self.timer.start()
    def drive_end(self):
        self.timer.stop()
        self.tab2_tab.setCurrentCell(0, 0)
        self.tab2_tab.scrollToItem(self.tab2_tab.item(0, 0))
    def save_data(self):
        dlg = QFileDialog()
        dlg.setFileMode(QFileDialog.AnyFile)
        dlg.setNameFilter("Excel文件 (*.xlsx *.xls *.xlsm *.xlsb);;所有文件 (*)")
        dlg.setFilter(QDir.Files | QDir.NoDotAndDotDot)
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M")
        dlg.setWindowTitle("保存指令文件")
        dlg.setFileMode(QFileDialog.AnyFile)
        dlg.setAcceptMode(QFileDialog.AcceptSave)
        dlg.setDefaultSuffix("xlsx")
        dlg.selectFile(f"{timestamp}.xlsx")

        def save_to_excel(path: str):
            rows = self.tab2_tab.rowCount()
            cols = self.tab2_tab.columnCount()
            headers = []
            for col in range(cols):
                headers.append(self.tab2_tab.horizontalHeaderItem(col).text())
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            for row in range(rows):
                row_data = []
                for col in range(cols):
                    if col == 2:
                        combo = self.tab2_tab.cellWidget(row, col)
                        if combo:
                            row_data.append(combo.currentText())
                        else:
                            item = self.tab2_tab.item(row, col)
                            row_data.append(item.text() if item else "")
                    else:
                        item = self.tab2_tab.item(row, col)
                        row_data.append(item.text() if item else "")
                ws.append(row_data)
            wb.save(path)
        if dlg.exec_():
            selected_files = dlg.selectedFiles()
            if selected_files:
                file_path = selected_files[0]
                save_to_excel(file_path)
    def load_data(self):
        dlg = QFileDialog()
        dlg.setFileMode(QFileDialog.AnyFile)
        dlg.setNameFilter("Excel文件 (*.xlsx *.xls *.xlsm *.xlsb);;所有文件 (*)")
        dlg.setFilter(QDir.Files | QDir.NoDotAndDotDot)
        dlg.setWindowTitle("打开指令文件")
        dlg.setFileMode(QFileDialog.ExistingFile)
        def load_from_excel(path: str):
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.tab2_tab.setRowCount(0)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue

                frame_str = "" if row[0] is None else str(row[0]).strip()
                drive_time_str = "" if row[1] is None else str(row[1]).strip()
                camera_enable_str = "" if row[2] is None else str(row[2]).strip()

                if not any([frame_str, drive_time_str, camera_enable_str]):
                    continue

                frame_list = []
                if frame_str:
                    try:
                        if frame_str.startswith('[') and frame_str.endswith(']'):
                            frame_list = ast.literal_eval(frame_str)
                        else:
                            frame_list = [int(x.strip()) for x in frame_str.split(',') if x.strip().isdigit()]
                    except:
                        pass
                drive_time = 1
                if drive_time_str:
                    try:
                        drive_time_val = int(drive_time_str)
                        if 1 <= drive_time_val < 5:
                            drive_time = drive_time_val
                    except:
                        pass
                camera_enable = "关"
                if camera_enable_str:
                    if camera_enable_str in ['开', '关']:
                        camera_enable = camera_enable_str
                self.add_one_row(frame_list=frame_list, drive_time=drive_time, 
                               camera_enable=camera_enable, insert_row=self.tab2_tab.rowCount())
        if dlg.exec_():
            selected_files = dlg.selectedFiles()
            if selected_files:
                file_path = selected_files[0]
                load_from_excel(file_path)
    def add_one_row(self, frame_list=None, drive_time=1, camera_enable="关", insert_row=None):
        """
        添加一行数据到表格
        :param frame_list: 帧列表，如[1,2,3]
        :param drive_time: 驱动时间，1<=int<5
        :param camera_enable: 是否开启摄像头，"开"或"关"
        :param insert_row: 插入的行位置，None表示在最后添加
        """
        if frame_list is None:
            frame_list = []
        if not (1 <= drive_time <= 5):
            drive_time = 1
        if camera_enable not in ["开", "关"]:
            camera_enable = "关"
        if insert_row is None:
            insert_row = self.tab2_tab.rowCount()

        self.tab2_tab.insertRow(insert_row)

        row_item = QTableWidgetItem(str(insert_row + 1))
        row_item.setTextAlignment(Qt.AlignCenter)
        self.tab2_tab.setVerticalHeaderItem(insert_row, row_item)

        if frame_list:
            frame_str = "[" + ",".join(str(x) for x in frame_list) + "]"
        else:
            frame_str = "[]"
        frame_item = QTableWidgetItem(frame_str)
        frame_item.setTextAlignment(Qt.AlignCenter)
        self.tab2_tab.setItem(insert_row, 0, frame_item)

        drive_time_item = QTableWidgetItem(str(drive_time))
        drive_time_item.setTextAlignment(Qt.AlignCenter)
        self.tab2_tab.setItem(insert_row, 1, drive_time_item)
        camera_combo = QComboBox()
        camera_combo.wheelEvent = lambda event: event.ignore()  # 禁用鼠标滚轮
        camera_combo.addItems(["开", "关"])
        camera_combo.setStyleSheet("QComboBox { text-align: center; }")
        camera_combo.setEditable(True)
        camera_combo.lineEdit().setReadOnly(True)
        camera_combo.lineEdit().setAlignment(Qt.AlignCenter)
        if camera_enable == "开":
            camera_combo.setCurrentIndex(0)
        else:
            camera_combo.setCurrentIndex(1)
        self.tab2_tab.setCellWidget(insert_row, 2, camera_combo)

        for i in range(self.tab2_tab.rowCount()):
            self.tab2_tab.verticalHeaderItem(i).setText(str(i + 1))
    """底层函数"""
    def timer_send_data(self):
        current_row = self.tab2_tab.currentRow()
        drive_time = 1
        try:
            frame_item = self.tab2_tab.item(current_row, 0)
            if frame_item and frame_item.text().strip():
                frame_str = frame_item.text().strip()
                try:
                    if frame_str.startswith('[') and frame_str.endswith(']'):
                        frame_list = ast.literal_eval(frame_str)
                    else:
                        frame_list = [int(x.strip()) for x in frame_str.split(',') if x.strip()]
                    if frame_list:
                        self.dmf_chip_scene._clear_all_electrode_drive_select()
                        for frame_id in frame_list:
                            self.dmf_chip_scene.set_electrode_state(frame_id, drive=True, selected=True)
                        drive_ids = list(set(frame_list) | set(self.dmf_chip_scene.pinned_ids()))
                        data = self.dmf_chip_scene.data_transfer(drive_ids)
                        if sum(data) != 0:
                            self.driver.drive_cmd6(data)
                except Exception as e:
                    self.log.debug(False,f"解析或发送帧数据失败: {e}")

            drive_time_item = self.tab2_tab.item(current_row, 1)
            if drive_time_item and drive_time_item.text().strip():
                try:
                    drive_time = int(drive_time_item.text().strip())
                    if drive_time < 1:
                        drive_time = 1
                except Exception as e:
                    self.log.debug(False,f"驱动时间解析失败，使用默认值1秒: {e}")
                    drive_time = 1

            camera_combo = self.tab2_tab.cellWidget(current_row, 2)
            if camera_combo:
                camera_status = camera_combo.currentText()
                if CAM_ENABLE and hasattr(self, 'camera_logic') and self.camera_logic:
                    try:
                        if camera_status == "开":
                            self.camera_logic.open_camera()
                            self.log.debug(True,"摄像头已开启")
                        elif camera_status == "关":
                            self.camera_logic.close_camera()
                            self.log.debug(True,"摄像头已关闭")
                    except Exception as e:
                        self.log.debug(True,f"摄像头操作失败: {e}")

        except Exception as e:
            self.log.debug(False,f"执行第 {current_row + 1} 行出错: {e}")

        if current_row >= self.tab2_tab.rowCount() - 1:
            self.timer.stop()
        else:
            self.timer.setInterval(drive_time * 1000)
            self.tab2_tab.setCurrentCell(current_row + 1, 0)
            self.tab2_tab.scrollToItem(self.tab2_tab.item(current_row + 1, 0))
    def set_voltage(self):
        if self.lineEdit1.text()=='':
            return None
        else:
            data = int(self.lineEdit1.text())
            if (data <= 300 and data >= 40):
                high_byte=(data>>8)&0xff
                low_byte=data&0xff
                self.driver.drive_cmd4([high_byte, low_byte])
                self.serial_information.append("Send:"+''.join(f'\\x{b:02x}' for b in self.driver.data))
    def set_acdc(self):
        if self.lineEdit2.text()=='':
            return None
        else:
            data = int(self.lineEdit2.text())
            if data <=2  or data >=0:
                self.driver.drive_cmd5(data)
                self.driver.voltage_module=data
                self.serial_information.append("Send:" + ''.join(f'\\x{b:02x}' for b in self.driver.data))
    def set_times(self):
        if self.lineEdit3_1.text()=='':
            return None
        else:
            data = int(self.lineEdit3_1.text())
            if data <=100  or data >=0:
                if data != 0:
                    self.driver.drive_cmd5(data)
                    self.serial_information.append("Send:" + ''.join(f'\\x{b:02x}' for b in self.driver.data))
    def set_times1(self):
        if self.lineEdit3_2.text()=='':
            return None
        else:
            data = int(self.lineEdit3_2.text())
            if data <=100  or data >=0:
                if data != 0:
                    self.driver.drive_cmd6(data)
                    self.serial_information.append("Send:" + ''.join(f'\\x{b:02x}' for b in self.driver.data))
    """液滴位置检测"""
    def detect_droplet(self):
        position=self.driver.drive_cmd3()  #读取控制点位
        if sum(position)!=0:
            positions = []
            for byte_index, value in enumerate(position):
                try:
                    v = int(value) & 0xFF
                except (TypeError, ValueError):
                    continue
                for bit_index in range(8):
                    if v & (1 << bit_index):
                        pos = byte_index * 8 + bit_index + 1
                        positions.append(pos)
            positions.sort()
            self.driver.drive_cmd5(40)  #设定检测电压值
            time.sleep(1) #延迟一秒等待电压修改完成
            temp_position_data=[]
            temp_value_data=[]
            # 采集：以 positions 为中心，取其周围3x3点位（与原逻辑一致）
            # 新规则：每个点位只读取1次ADC
            for pos_index in positions:
                xy = self.dmf_chip_scene.id_position.get(pos_index)
                if not xy:
                    continue
                x, y = xy
                new_index = [x - 1, y - 1]
                for i in range(0, 3):
                    for j in range(0, 3):
                        try:
                            value = self.dmf_chip_scene.id_array[new_index[0] + i][new_index[1] + j]
                        except Exception:
                            continue
                        if not value:
                            continue
                        # 打开点位
                        self.driver.drive_cmd6(self.dmf_chip_scene.data_transfer([value]))
                        time.sleep(0.3)  # 等待点位稳定
                        adc_val = self.driver.bottom_cmd2()
                        if adc_val is None:
                            continue
                        try:
                            adc_val = float(adc_val)
                        except (TypeError, ValueError):
                            continue
                        temp_position_data.append(value)
                        temp_value_data.append(adc_val)

            if not temp_position_data or not temp_value_data:
                self.log.debug(False, "液滴检测：未获取到有效ADC数据")
                return

            # 概率规则：
            # - 参考ADC值对应概率=1
            # - 偏离参考最远的点概率=0
            # - 其它点按偏离程度线性分布在 0.1~0.9
            # 默认参考点取第一组3x3的中心（索引4）；不足则取第一个
            ref_idx = 4 if len(temp_value_data) > 4 else 0
            ref_val = temp_value_data[ref_idx]

            devs = [abs(v - ref_val) for v in temp_value_data]
            max_dev = max(devs) if devs else 0.0

            probs = []
            for idx, dev in enumerate(devs):
                if idx == ref_idx:
                    probs.append(1.0)
                    continue
                if max_dev <= 0:
                    # 全部等于参考：其它点给 0.9
                    probs.append(0.9)
                    continue
                if dev == max_dev:
                    probs.append(0.0)
                    continue
                p = 0.9 - 0.8 * (dev / max_dev)
                if p < 0.1:
                    p = 0.1
                if p > 0.9:
                    p = 0.9
                probs.append(p)
            for pos, p in zip(temp_position_data, probs):
                self.add_droplet_segment(pos, p)
    def add_droplet_segment(self, x_value, y_value):
        if not hasattr(self, "droplet_x") or not hasattr(self, "droplet_y"):
            return
        if not hasattr(self, "droplet_segment_count"):
            self.droplet_segment_count = 0
        if not hasattr(self, "droplet_text_items"):
            self.droplet_text_items = []

        if self.droplet_segment_count >= 9:
            self.clear_droplet_plot()
        try:
            x_int = int(round(float(x_value)))
        except (TypeError, ValueError):
            return
        if x_int < 1 or x_int > len(self.droplet_x):
            return
        try:
            y_float = float(y_value)
        except (TypeError, ValueError):
            return
        if y_float < 0.0:
            y_float = 0.0
        if y_float > 1.0:
            y_float = 1.0
        y_float = round(y_float, 2)
        self.droplet_y[x_int] = y_float
        self.droplet_curve.setData(self.droplet_x, self.droplet_y)
        plot_item = self.droplet_plot_widget.getPlotItem()
        text = str(x_int)
        text_x = x_int
        text_y = y_float + 0.07
        text_item = pg.TextItem(text, anchor=(0.5, 0),color='k')
        text_item.setPos(text_x, text_y)
        plot_item.addItem(text_item)
        self.droplet_text_items.append(text_item)
        self.droplet_segment_count += 1
    def clear_droplet_plot(self):
        if not hasattr(self, "droplet_x") or not hasattr(self, "droplet_y"):
            return
        plot_item = self.droplet_plot_widget.getPlotItem()
        if hasattr(self, "droplet_text_items"):
            for item in self.droplet_text_items:
                plot_item.removeItem(item)
            self.droplet_text_items = []  #
        self.droplet_y = [0.0] * len(self.droplet_y)
        self.droplet_curve.setData(self.droplet_x, self.droplet_y)
        self.droplet_segment_count = 0
    """温度读取和绘制相关函数"""
    def read_temperature_and_plot(self, channel):
        """读取指定通道的温度数据并更新曲线

        Args:
            channel: 通道号 (1, 2, 3)
        """
        try:
            if channel == 1:
                # 读取通道1（左侧）的温度和功率
                temp = self.driver.temp_cmd7()/10
                time.sleep(0.01)
                power = self.driver.temp_cmdb()
                if power is not None:
                    power = float(power)/10
                if temp is not None:
                    self.update_temp_curve(1, temp, power)
            elif channel == 2:
                # 读取通道2（右侧）的温度和功率
                temp = self.driver.temp_cmd8()/10
                time.sleep(0.01)
                power = self.driver.temp_cmdc()
                if power is not None:
                    power = float(power)/10
                if temp is not None:
                    self.update_temp_curve(2, temp, power)
            elif channel == 3:
                # 通道3暂时使用通道2的数据（如果有需要可以单独实现）
                temp = self.driver.temp_cmd8()/10
                time.sleep(0.01)
                power = self.driver.temp_cmdc()
                if power is not None:
                    power = float(power)/10
                if temp is not None:
                    self.update_temp_curve(3, temp, power)
            
        except Exception as e:
            self.log.debug(False, f"读取通道{channel}温度数据失败: {e}")
    def update_temp_curve(self, channel, temperature, power=None):
        """更新指定通道的温度曲线
        
        Args:
            channel: 通道号 (1, 2, 3)
            temperature: 温度值
            power: 功率值（可选）
        """
        # 获取该通道的开始时间
        start_time = self.temp_start_time.get(channel, time.time())
        current_time = time.time() - start_time  # 相对时间（秒）
        
        if channel == 1:
            self.temp1_time.append(current_time)
            self.temp1_data.append(temperature)
            if power is not None:
                self.power1_data.append(float(power))
            
            # 限制数据点数量
            if len(self.temp1_time) > self.max_data_points:
                self.temp1_time = self.temp1_time[-self.max_data_points:]
                self.temp1_data = self.temp1_data[-self.max_data_points:]
                if len(self.power1_data) > self.max_data_points:
                    self.power1_data = self.power1_data[-self.max_data_points:]
            
            # 更新曲线
            self.temperature_curve1.setData(self.temp1_time, self.temp1_data)
            if power is not None and len(self.power1_data) > 0:
                self.power_curve1.setData(self.temp1_time, self.power1_data)
            
            # 自动调整视图范围以最佳显示
            if len(self.temp1_time) > 0:
                plot_item1 = self.temperature_widget1.getPlotItem()
                plot_item1.getViewBox().autoRange(padding=0.02)  # 2%的边距
                
        elif channel == 2:
            self.temp2_time.append(current_time)
            self.temp2_data.append(temperature)
            if power is not None:
                self.power2_data.append(float(power))
            
            # 限制数据点数量
            if len(self.temp2_time) > self.max_data_points:
                self.temp2_time = self.temp2_time[-self.max_data_points:]
                self.temp2_data = self.temp2_data[-self.max_data_points:]
                if len(self.power2_data) > self.max_data_points:
                    self.power2_data = self.power2_data[-self.max_data_points:]
            
            # 更新曲线
            self.temperature_curve2.setData(self.temp2_time, self.temp2_data)
            if power is not None and len(self.power2_data) > 0:
                self.power_curve2.setData(self.temp2_time, self.power2_data)
            
            # 自动调整视图范围以最佳显示
            if len(self.temp2_time) > 0:
                plot_item2 = self.temperature_widget2.getPlotItem()
                plot_item2.getViewBox().autoRange(padding=0.02)  # 2%的边距
                
        elif channel == 3:
            self.temp3_time.append(current_time)
            self.temp3_data.append(temperature)
            if power is not None:
                self.power3_data.append(float(power))
            
            # 限制数据点数量
            if len(self.temp3_time) > self.max_data_points:
                self.temp3_time = self.temp3_time[-self.max_data_points:]
                self.temp3_data = self.temp3_data[-self.max_data_points:]
                if len(self.power3_data) > self.max_data_points:
                    self.power3_data = self.power3_data[-self.max_data_points:]
            
            # 更新曲线
            self.temperature_curve3.setData(self.temp3_time, self.temp3_data)
            if power is not None and len(self.power3_data) > 0:
                self.power_curve3.setData(self.temp3_time, self.power3_data)
            
            # 自动调整视图范围以最佳显示
            if len(self.temp3_time) > 0:
                plot_item3 = self.temperature_widget3.getPlotItem()
                plot_item3.getViewBox().autoRange(padding=0.02)  # 2%的边距
    def start_temp_monitoring(self, channel):
        """开始温度监测

        流程：
        1. 读取设定温度并发送设定温度指令
        2. 发送开始指令
        3. 启动定时器开始读取温度功率（100ms间隔，1秒10次）

        Args:
            channel: 通道号 (1, 2, 3)
        """
        try:
            # 1. 读取设定温度
            if channel == 1:
                target_temp_str = self.lineEdit_target_temp1.text().strip()
                if not target_temp_str:
                    self.log.debug(False, f"通道{channel}：请先输入设定温度")
                    return
                try:
                    target_temp = int(target_temp_str)*10
                    # 转换为高字节和低字节（高字节前，低字节后）
                    temp_int = int(target_temp)  # 假设温度
                    high_byte = (temp_int >> 8) & 0xFF
                    low_byte = temp_int & 0xFF
                    # 发送设定温度指令（左侧）
                    self.driver.temp_cmd3([high_byte, low_byte])
                    self.log.debug(True, f"通道{channel}：设定温度 {target_temp}°C")
                except ValueError:
                    self.log.debug(False, f"通道{channel}：设定温度格式错误")
                    return
                # 发送开始指令（左侧）
                result = self.driver.temp_cmd1(0x01)  # 1表示启动
                if result == 1:
                    self.log.debug(True, f"通道{channel}：启动指令发送成功")
                else:
                    self.log.debug(False, f"通道{channel}：启动指令发送失败")
            elif channel == 2:
                target_temp_str = self.lineEdit_target_temp2.text().strip()
                if not target_temp_str:
                    self.log.debug(False, f"通道{channel}：请先输入设定温度")
                    return
                try:
                    target_temp = float(target_temp_str)
                    temp_int = int(target_temp)*10
                    high_byte = (temp_int >> 8) & 0xFF
                    low_byte = temp_int & 0xFF
                    # 发送设定温度指令（右侧）
                    self.driver.temp_cmd4([high_byte, low_byte])
                    self.log.debug(True, f"通道{channel}：设定温度 {target_temp}°C")
                except ValueError:
                    self.log.debug(False, f"通道{channel}：设定温度格式错误")
                    return
                # 发送开始指令（右侧）
                result = self.driver.temp_cmd2(0x1)  # 1表示启动
                if result == 1:
                    self.log.debug(True, f"通道{channel}：启动指令发送成功")
                else:
                    self.log.debug(False, f"通道{channel}：启动指令发送失败")
            elif channel == 3:
                # 通道3暂时使用通道2的逻辑
                target_temp_str = self.lineEdit_target_temp3.text().strip()
                if not target_temp_str:
                    self.log.debug(False, f"通道{channel}：请先输入设定温度")
                    return
                try:
                    target_temp = float(target_temp_str)
                    temp_int = int(target_temp)*10
                    high_byte = (temp_int >> 8) & 0xFF
                    low_byte = temp_int & 0xFF
                    # 通道3暂时使用通道2的指令
                    self.driver.temp_cmd4([high_byte, low_byte])
                    self.log.debug(True, f"通道{channel}：设定温度 {target_temp}°C")
                except ValueError:
                    self.log.debug(False, f"通道{channel}：设定温度格式错误")
                    return
                result = self.driver.temp_cmd2(0x01)
                if result == 1:
                    self.log.debug(True, f"通道{channel}：启动指令发送成功")
                else:
                    self.log.debug(False, f"通道{channel}：启动指令发送失败")
            
            # 2. 启动定时器开始读取
            if channel in self.temp_timers:
                if not self.temp_timers[channel].isActive():
                    self.temp_start_time[channel] = time.time()  # 记录开始时间
                    self.temp_timers[channel].start()
                    self.log.debug(True, f"通道{channel}：开始监测温度（100ms间隔）")
                else:
                    self.log.debug(True, f"通道{channel}：温度监测已在运行中")
                    
        except Exception as e:
            self.log.debug(False, f"通道{channel}启动失败: {e}")
    def pause_temp_monitoring(self, channel):
        """暂停温度监测

        流程：
        1. 发送关闭指令
        2. 停止定时器，不再读取数据

        Args:
            channel: 通道号 (1, 2, 3)
        """
        try:
            # 1. 发送关闭指令
            if channel == 1:
                result = self.driver.temp_cmd1(0x00)  # 0表示停止
                if result == 0:
                    self.log.debug(True, f"通道{channel}：关闭指令发送成功")
            elif channel == 2:
                result = self.driver.temp_cmd2(0)  # 0表示停止
                if result == 0:
                    self.log.debug(True, f"通道{channel}：关闭指令发送成功")
            elif channel == 3:
                # 通道3暂时使用通道2的指令
                result = self.driver.temp_cmd2(0)
                if result == 0:
                    self.log.debug(True, f"通道{channel}：关闭指令发送成功")
            
            # 2. 停止定时器
            if channel in self.temp_timers:
                if self.temp_timers[channel].isActive():
                    self.temp_timers[channel].stop()
                    self.log.debug(True, f"通道{channel}：已停止监测温度")
                else:
                    self.log.debug(True, f"通道{channel}：温度监测未在运行")
                    
        except Exception as e:
            self.log.debug(False, f"通道{channel}暂停失败: {e}")
    def send_pid_params(self, channel):
        try:
            if channel == 1:
                p_str = self.lineEdit_P1.text().strip()
                i_str = self.lineEdit_I1.text().strip()
                d_str = self.lineEdit_D1.text().strip()
                try:
                    p_value = int(p_str) if p_str else 0
                    i_value = int(i_str) if i_str else 0
                    d_value = int(d_str) if d_str else 0
                except ValueError:
                    self.log.debug(False, f"通道{channel}：PID参数格式错误，请输入整数")
                    return

                pid_bytes = []
                for value in [p_value, i_value, d_value]:
                    high_byte = (value >> 8) & 0xFF
                    low_byte = value & 0xFF
                    pid_bytes.append(high_byte)
                    pid_bytes.append(low_byte)

                result = self.driver.temp_cmd5(pid_bytes)
                if result:
                    self.log.debug(True, f"通道{channel}：PID参数发送成功 P={p_value}, I={i_value}, D={d_value}")
                else:
                    self.log.debug(False, f"通道{channel}：PID参数发送失败")
            elif channel == 2:
                p_str = self.lineEdit_P2.text().strip()
                i_str = self.lineEdit_I2.text().strip()
                d_str = self.lineEdit_D2.text().strip()
                try:
                    p_value = int(p_str) if p_str else 0
                    i_value = int(i_str) if i_str else 0
                    d_value = int(d_str) if d_str else 0
                except ValueError:
                    self.log.debug(False, f"通道{channel}：PID参数格式错误，请输入整数")
                    return
                pid_bytes = []
                for value in [p_value, i_value, d_value]:
                    high_byte = (value >> 8) & 0xFF
                    low_byte = value & 0xFF
                    pid_bytes.append(high_byte)
                    pid_bytes.append(low_byte)
                
                # 发送PID参数设定指令（右侧）
                result = self.driver.temp_cmd6(pid_bytes)
                if result:
                    self.log.debug(True, f"通道{channel}：PID参数发送成功 P={p_value}, I={i_value}, D={d_value}")
                else:
                    self.log.debug(False, f"通道{channel}：PID参数发送失败")
                    
            elif channel == 3:
                # 通道3暂时使用通道2的指令
                p_str = self.lineEdit_P3.text().strip()
                i_str = self.lineEdit_I3.text().strip()
                d_str = self.lineEdit_D3.text().strip()
                
                try:
                    p_value = int(p_str) if p_str else 0
                    i_value = int(i_str) if i_str else 0
                    d_value = int(d_str) if d_str else 0
                except ValueError:
                    self.log.debug(False, f"通道{channel}：PID参数格式错误，请输入整数")
                    return
                
                pid_bytes = []
                for value in [p_value, i_value, d_value]:
                    high_byte = (value >> 8) & 0xFF
                    low_byte = value & 0xFF
                    pid_bytes.append(high_byte)
                    pid_bytes.append(low_byte)
                
                # 通道3暂时使用通道2的指令
                result = self.driver.temp_cmd6(pid_bytes)
                if result:
                    self.log.debug(True, f"通道{channel}：PID参数发送成功 P={p_value}, I={i_value}, D={d_value}")
                else:
                    self.log.debug(False, f"通道{channel}：PID参数发送失败")
                    
        except Exception as e:
            self.log.debug(False, f"通道{channel}发送PID参数失败: {e}")
    def clear_temp_data(self, channel=None):
        """清除温度数据

        Args:
            channel: 通道号 (1, 2, 3)，如果为None则清除所有通道
        """
        if channel is None or channel == 1:
            self.temp1_time = []
            self.temp1_data = []
            self.power1_data = []
            self.temperature_curve1.setData([], [])
            self.power_curve1.setData([], [])
            if 1 in self.temp_start_time:
                del self.temp_start_time[1]
        
        if channel is None or channel == 2:
            self.temp2_time = []
            self.temp2_data = []
            self.power2_data = []
            self.temperature_curve2.setData([], [])
            self.power_curve2.setData([], [])
            if 2 in self.temp_start_time:
                del self.temp_start_time[2]
        
        if channel is None or channel == 3:
            self.temp3_time = []
            self.temp3_data = []
            self.power3_data = []
            self.temperature_curve3.setData([], [])
            self.power_curve3.setData([], [])
            if 3 in self.temp_start_time:
                del self.temp_start_time[3]
        
        self.log.debug(True, f"已清除通道{channel if channel else '所有'}的温度数据")
    def save_temp_data(self, channel):
        """保存指定通道的温度数据到Excel文件

        Args:
            channel: 通道号 (1, 2, 3)
        """
        try:
            # 获取当前通道的数据
            if channel == 1:
                time_data = self.temp1_time
                temp_data = self.temp1_data
                power_data = self.power1_data
            elif channel == 2:
                time_data = self.temp2_time
                temp_data = self.temp2_data
                power_data = self.power2_data
            elif channel == 3:
                time_data = self.temp3_time
                temp_data = self.temp3_data
                power_data = self.power3_data
            else:
                self.log.debug(False, f"无效的通道号: {channel}")
                return
            
            # 检查是否有数据
            if not time_data or not temp_data:
                self.log.debug(False, f"通道{channel}：没有数据可保存")
                return
            
            # 打开文件保存对话框
            dlg = QFileDialog()
            dlg.setFileMode(QFileDialog.AnyFile)
            dlg.setNameFilter("Excel文件 (*.xlsx *.xls);;所有文件 (*)")
            dlg.setFilter(QDir.Files | QDir.NoDotAndDotDot)
            timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
            dlg.setWindowTitle(f"保存通道{channel}温度数据")
            dlg.setFileMode(QFileDialog.AnyFile)
            dlg.setAcceptMode(QFileDialog.AcceptSave)
            dlg.setDefaultSuffix("xlsx")
            dlg.selectFile(f"通道{channel}_温度数据_{timestamp}.xlsx")
            
            if dlg.exec_():
                selected_files = dlg.selectedFiles()
                if selected_files:
                    file_path = selected_files[0]
                    
                    # 准备数据
                    # 确保时间、温度、功率数据长度一致
                    min_len = min(len(time_data), len(temp_data))
                    if power_data:
                        min_len = min(min_len, len(power_data))
                    
                    # 使用openpyxl保存为Excel
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = '温度数据'
                    
                    # 写入表头
                    headers = ['时间(s)', '温度(℃)']
                    power_col_idx = None
                    temp_col_idx = 2  # 温度列的索引（1-based，列B是2）
                    if power_data and len(power_data) >= min_len:
                        headers.append('功率(W)')
                        power_col_idx = len(headers)  # 功率列的索引（1-based，列C是3）
                    ws.append(headers)
                    
                    # 写入数据
                    for i in range(min_len):
                        row_data = [time_data[i], temp_data[i]]
                        if power_data and len(power_data) >= min_len:
                            power_val = float(power_data[i])
                            row_data.append(power_val)
                        ws.append(row_data)
                    
                    # 设置温度列的数字格式，确保小数能正确显示
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=temp_col_idx, max_col=temp_col_idx):
                        for cell in row:
                            cell.number_format = '0.00'  # 设置为保留2位小数的格式
                    
                    # 设置功率列的数字格式，确保小数能正确显示
                    if power_col_idx:
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=power_col_idx, max_col=power_col_idx):
                            for cell in row:
                                cell.number_format = '0.00'  # 设置为保留2位小数的格式
                    
                    # 保存文件
                    wb.save(file_path)
                    
                    self.log.debug(True, f"通道{channel}：温度数据已保存到 {file_path}")
                    
        except Exception as e:
            self.log.debug(False, f"通道{channel}保存数据失败: {e}")
    def load_temp_data(self, channel):
        try:
            dlg = QFileDialog()
            dlg.setFileMode(QFileDialog.AnyFile)
            dlg.setNameFilter("Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;所有文件 (*)")
            dlg.setFilter(QDir.Files | QDir.NoDotAndDotDot)
            dlg.setWindowTitle(f"加载通道{channel}温度数据")
            dlg.setFileMode(QFileDialog.ExistingFile)
            
            if dlg.exec_():
                selected_files = dlg.selectedFiles()
                if selected_files:
                    file_path = selected_files[0]
                    
                    # 读取文件（支持Excel和CSV格式，向后兼容）
                    try:
                        if file_path.lower().endswith(('.xlsx', '.xls')):
                            from openpyxl import load_workbook
                            wb = load_workbook(file_path, data_only=True)
                            ws = wb.active

                            headers = []
                            for cell in ws[1]:
                                headers.append(cell.value if cell.value else '')

                            if '时间(s)' not in headers or '温度(℃)' not in headers:
                                self.log.debug(False, f"通道{channel}：Excel文件格式错误，缺少必需的列")
                                return

                            time_col_idx = headers.index('时间(s)') + 1
                            temp_col_idx = headers.index('温度(℃)') + 1
                            power_col_idx = headers.index('功率(W)') + 1 if '功率(W)' in headers else None

                            time_data = []
                            temp_data = []
                            power_data = []
                            
                            for row in ws.iter_rows(min_row=2, values_only=False):
                                time_val = row[time_col_idx - 1].value
                                if time_val is not None:
                                    try:
                                        time_data.append(float(time_val))
                                    except (ValueError, TypeError):
                                        continue

                                temp_val = row[temp_col_idx - 1].value
                                if temp_val is not None:
                                    try:
                                        temp_data.append(float(temp_val))
                                    except (ValueError, TypeError):
                                        continue

                                if power_col_idx:
                                    power_val = row[power_col_idx - 1].value
                                    if power_val is not None:
                                        try:
                                            power_data.append(float(power_val))
                                        except (ValueError, TypeError):
                                            power_data.append(0.0)
                                    else:
                                        power_data.append(0.0)

                            min_len = min(len(time_data), len(temp_data))
                            if power_data:
                                min_len = min(min_len, len(power_data))
                                power_data = power_data[:min_len]
                            time_data = time_data[:min_len]
                            temp_data = temp_data[:min_len]

                            if not power_data or len(power_data) == 0:
                                power_data = None
                                
                        else:
                            df = pd.read_csv(file_path, encoding='utf-8-sig')

                            if '时间(s)' not in df.columns or '温度(℃)' not in df.columns:
                                self.log.debug(False, f"通道{channel}：CSV文件格式错误，缺少必需的列")
                                return

                            time_data = df['时间(s)'].tolist()
                            temp_data = df['温度(℃)'].tolist()

                            power_data = None
                            if '功率(W)' in df.columns:
                                power_data = df['功率(W)'].tolist()

                            time_data = [float(x) for x in time_data if pd.notna(x)]
                            temp_data = [float(x) for x in temp_data if pd.notna(x)]
                            if power_data:
                                power_data = [float(x) if pd.notna(x) else 0.0 for x in power_data]
                                
                    except Exception as e:
                        self.log.debug(False, f"通道{channel}：文件格式错误，无法读取: {e}")
                        return

                    if channel == 1:
                        self.temp1_time = time_data
                        self.temp1_data = temp_data
                        if power_data:
                            self.power1_data = power_data
                        else:
                            self.power1_data = []

                        self.temperature_curve1.setData(self.temp1_time, self.temp1_data)
                        if power_data and len(power_data) > 0:
                            self.power_curve1.setData(self.temp1_time, self.power1_data)
                        else:
                            self.power_curve1.setData([], [])

                        if len(self.temp1_time) > 0:
                            plot_item1 = self.temperature_widget1.getPlotItem()
                            # 调整温度ViewBox的范围（X轴和Y轴）
                            plot_item1.getViewBox().autoRange(padding=0.02)
                            # 功率轴已设置为固定范围0-5，不需要自动调整
                        
                    elif channel == 2:
                        self.temp2_time = time_data
                        self.temp2_data = temp_data
                        if power_data:
                            self.power2_data = power_data
                        else:
                            self.power2_data = []

                        self.temperature_curve2.setData(self.temp2_time, self.temp2_data)
                        if power_data and len(power_data) > 0:
                            self.power_curve2.setData(self.temp2_time, self.power2_data)
                        else:
                            self.power_curve2.setData([], [])

                        if len(self.temp2_time) > 0:
                            plot_item2 = self.temperature_widget2.getPlotItem()
                            # 调整温度ViewBox的范围（X轴和Y轴）
                            plot_item2.getViewBox().autoRange(padding=0.02)
                            # 功率轴已设置为固定范围0-5，不需要自动调整
                        
                    elif channel == 3:
                        self.temp3_time = time_data
                        self.temp3_data = temp_data
                        if power_data:
                            self.power3_data = power_data
                        else:
                            self.power3_data = []

                        self.temperature_curve3.setData(self.temp3_time, self.temp3_data)
                        if power_data and len(power_data) > 0:
                            self.power_curve3.setData(self.temp3_time, self.power3_data)
                        else:
                            self.power_curve3.setData([], [])

                        if len(self.temp3_time) > 0:
                            plot_item3 = self.temperature_widget3.getPlotItem()
                            # 调整温度ViewBox的范围（X轴和Y轴）
                            plot_item3.getViewBox().autoRange(padding=0.02)
                            # 功率轴已设置为固定范围0-5，不需要自动调整
                    
                    self.log.debug(True, f"通道{channel}：已加载 {len(time_data)} 个数据点")
                    
        except Exception as e:
            self.log.debug(False, f"通道{channel}加载数据失败: {e}")

if __name__ == "__main__":
    import sys
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    mywindow = MainWindow()
    mywindow.show()
    sys.exit(app.exec_())
